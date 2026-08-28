"""استيراد وتصدير أعطال من/إلى Excel — قالب وتصدير بترويسة احترافية."""

from __future__ import annotations

from datetime import datetime, time

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from webapp import db
from webapp import excel_brand as brand

TICKET_HEADERS = [
    "رقم العطل",
    "كود ركاز",
    "تاريخ الاستلام",
    "الحي",
    "وقت الاستلام",
    "المندوب",
    "رقم المحطة",
    "رقم الفيدر",
    "الموقع",
    "نوع العطل",
    "تصنيف العطل",
    "الفرقة",
    "وقت التوجيه",
    "وقت الوصول",
    "حالة التنفيذ",
    "تاريخ التنفيذ",
    "تم التصوير",
    "الكميات مكتملة",
    "إخلاء الأسفلت",
    "حالة التمتير",
    "اعتماد الاستشاري",
    "حالة المستخلص",
    "أمر العمل",
    "رقم الفاتورة",
    "حالة SAP",
    "قيمة البنود",
    "ملاحظات",
]

TICKET_FIELDS = [
    "ticket_no",
    "rekaz_code",
    "receive_date",
    "district",
    "receive_time",
    "agent",
    "station_no",
    "feeder_no",
    "location",
    "fault_type",
    "classification",
    "team",
    "dispatch_time",
    "arrival_time",
    "status",
    "execution_date",
    "photographed",
    "quantities_done",
    "asphalt_clearance",
    "metering_status",
    "consultant_approval",
    "invoice_status",
    "work_order",
    "invoice_no",
    "sap_status",
    "items_value",
    "notes",
]

# أعمدة التصدير الكامل = نفس قالب الاستيراد (للتوافق وإعادة الاستيراد)
EXPORT_HEADERS = TICKET_HEADERS
EXPORT_FIELDS = TICKET_FIELDS

_COL_WIDTHS = {
    "رقم العطل": 14,
    "كود ركاز": 12,
    "تاريخ الاستلام": 14,
    "الحي": 12,
    "وقت الاستلام": 12,
    "المندوب": 14,
    "رقم المحطة": 12,
    "رقم الفيدر": 12,
    "الموقع": 28,
    "نوع العطل": 16,
    "تصنيف العطل": 14,
    "الفرقة": 12,
    "وقت التوجيه": 12,
    "وقت الوصول": 12,
    "حالة التنفيذ": 14,
    "تاريخ التنفيذ": 14,
    "تم التصوير": 12,
    "الكميات مكتملة": 14,
    "إخلاء الأسفلت": 14,
    "حالة التمتير": 14,
    "اعتماد الاستشاري": 16,
    "حالة المستخلص": 14,
    "أمر العمل": 14,
    "رقم الفاتورة": 14,
    "حالة SAP": 12,
    "قيمة البنود": 12,
    "ملاحظات": 28,
}

_DATE_FIELDS = {"receive_date", "execution_date"}
_TIME_FIELDS = {"receive_time", "dispatch_time", "arrival_time"}

_TICKET_ALIASES = {
    "رقم العطل": "ticket_no",
    "رقم البلاغ": "ticket_no",  # توافق مع القوالب القديمة
    "البلاغ": "ticket_no",
    "ticket_no": "ticket_no",
    "ticket no": "ticket_no",
    "fault no": "ticket_no",
    "fault_no": "ticket_no",
    "كود ركاز": "rekaz_code",
    "كود er": "rekaz_code",
    "كود ER": "rekaz_code",
    "رقم ركاز": "rekaz_code",
    "rekaz_code": "rekaz_code",
    "er code": "rekaz_code",
    "تاريخ الاستلام": "receive_date",
    "تاريخ العطل": "receive_date",
    "تاريخ البلاغ": "receive_date",  # توافق
    "receive_date": "receive_date",
    "الحي": "district",
    "district": "district",
    "وقت الاستلام": "receive_time",
    "receive_time": "receive_time",
    "المندوب": "agent",
    "agent": "agent",
    "رقم المحطة": "station_no",
    "المحطة": "station_no",
    "station_no": "station_no",
    "رقم الفيدر": "feeder_no",
    "الفيدر": "feeder_no",
    "feeder_no": "feeder_no",
    "الموقع": "location",
    "الموقع (رابط خرائط)": "location",
    "location": "location",
    "نوع العطل": "fault_type",
    "العطل": "fault_type",
    "fault_type": "fault_type",
    "تصنيف العطل": "classification",
    "تصنيف البلاغ": "classification",  # توافق
    "التصنيف": "classification",
    "classification": "classification",
    "الفرقة": "team",
    "team": "team",
    "وقت التوجيه": "dispatch_time",
    "dispatch_time": "dispatch_time",
    "وقت الوصول": "arrival_time",
    "arrival_time": "arrival_time",
    "حالة التنفيذ": "status",
    "الحالة": "status",
    "status": "status",
    "تاريخ التنفيذ": "execution_date",
    "execution_date": "execution_date",
    "تم التصوير": "photographed",
    "التصوير": "photographed",
    "photographed": "photographed",
    "الكميات مكتملة": "quantities_done",
    "quantities_done": "quantities_done",
    "إخلاء الأسفلت": "asphalt_clearance",
    "asphalt_clearance": "asphalt_clearance",
    "حالة التمتير": "metering_status",
    "metering_status": "metering_status",
    "اعتماد الاستشاري": "consultant_approval",
    "consultant_approval": "consultant_approval",
    "حالة المستخلص": "invoice_status",
    "invoice_status": "invoice_status",
    "أمر العمل": "work_order",
    "رقم أمر العمل": "work_order",
    "رقم المعاملة": "work_order",  # توافق مع التسمية القديمة
    "work_order": "work_order",
    "work order no": "work_order",
    "work order number": "work_order",
    "transaction no": "work_order",
    "transaction number": "work_order",
    "رقم الفاتورة": "invoice_no",
    "invoice_no": "invoice_no",
    "حالة sap": "sap_status",
    "حالة SAP": "sap_status",
    "sap_status": "sap_status",
    "قيمة البنود": "items_value",
    "items_value": "items_value",
    "ملاحظات": "notes",
    "notes": "notes",
}


def _norm_header(value) -> str:
    return str(value or "").strip().lower().replace("ـ", "")


def _map_headers(row_values) -> dict[int, str]:
    mapping = {}
    for idx, raw in enumerate(row_values):
        key = str(raw or "").strip()
        field = _TICKET_ALIASES.get(key) or _TICKET_ALIASES.get(key.lower())
        if not field:
            field = _TICKET_ALIASES.get(_norm_header(key))
        if field:
            mapping[idx] = field
    return mapping


def _find_header(rows, max_scan: int = 20) -> tuple[int, dict[int, str]]:
    """يعثر على صف رؤوس الأعمدة حتى مع وجود ترويسة/شعار أعلاه."""
    for i, row in enumerate(rows[:max_scan]):
        if not row:
            continue
        mapping = _map_headers(row)
        if "ticket_no" in mapping.values():
            return i, mapping
    return -1, {}


def _format_date(val) -> str:
    if val is None or val == "":
        return ""
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    if hasattr(val, "strftime") and not isinstance(val, time):
        try:
            return val.strftime("%Y-%m-%d")
        except Exception:
            pass
    s = str(val).strip()
    if " " in s and len(s) >= 10:
        s = s.split(" ")[0]
    if "/" in s and len(s.split("/")) == 3:
        parts = s.split("/")
        if len(parts[2]) == 4:
            d, m, y = parts[0].zfill(2), parts[1].zfill(2), parts[2]
            return f"{y}-{m}-{d}"
    return s


def _format_time(val) -> str:
    if val is None or val == "":
        return ""
    if isinstance(val, datetime):
        return val.strftime("%H:%M")
    if isinstance(val, time):
        return val.strftime("%H:%M")
    s = str(val).strip()
    if " " in s and ":" in s:
        part = s.split(" ")[-1]
        return part[:5] if len(part) >= 5 else part
    if len(s) >= 5 and s[2] == ":":
        return s[:5]
    return s


def _to_float(val):
    if val is None or val == "":
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace(",", "")
    if not s:
        return None
    try:
        return float(s)
    except Exception:
        return None


def _cell(row, idx, field: str | None = None):
    if idx is None or idx >= len(row):
        return ""
    val = row[idx]
    if val is None:
        return ""
    if field in _DATE_FIELDS:
        return _format_date(val)
    if field in _TIME_FIELDS:
        return _format_time(val)
    if field == "items_value":
        return _to_float(val)
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    return str(val).strip()


def _sample_row() -> list:
    today = datetime.now().strftime("%Y-%m-%d")
    return [
        "T-1001",
        today,
        "خريص",
        "08:30",
        "مندوب 1",
        "ST-01",
        "F-12",
        "",
        "عطل كيبل",
        "طارئ",
        "فرقة 1",
        "09:00",
        "09:25",
        "تم الإسناد",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        0,
        "صف مثال — احذفه أو عدّله",
    ]


def build_tickets_template() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "الأعطال"
    ncol = len(TICKET_HEADERS)

    header_row = brand.apply_brand_header(
        ws,
        title="قالب استيراد الأعطال",
        ncol=ncol,
    )
    brand.write_header_row(ws, TICKET_HEADERS, header_row, widths=_COL_WIDTHS)

    sample = _sample_row()
    data_row = header_row + 1
    for col, val in enumerate(sample, start=1):
        ws.cell(row=data_row, column=col, value=val)
    brand.style_data_rows(ws, start_row=data_row, end_row=data_row, ncol=ncol)
    # توسيع نطاق الفلتر ليشمل صف المثال
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(ncol)}{data_row}"

    brand.write_instructions_sheet(
        wb,
        [
            "عمود رقم العطل مطلوب وفريد.",
            "عند تكرار رقم العطل يتم تحديث السجل الموجود (upsert).",
            "التواريخ بصيغة YYYY-MM-DD والأوقات HH:MM.",
            "لا تحذف صف رؤوس الأعمدة ولا تغيّر أسماء الأعمدة.",
            "صف المثال يمكن حذفه أو تعديله قبل الرفع.",
            "ارفع الملف من صفحة الأعطال ← استيراد من Excel.",
        ],
    )
    return brand.save_workbook_bytes(wb)


def export_tickets(rows: list[dict] | None = None, title: str | None = None, filters: list[str] | None = None) -> bytes:
    """تصدير الأعطال بترويسة احترافية وشعار — أعمدة مطابقة للقالب."""
    if rows is None:
        conn = db.connect()
        rows = db.rows_to_dicts(conn.execute("SELECT * FROM tickets ORDER BY id").fetchall())
        conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "الأعطال"
    ncol = len(EXPORT_HEADERS)

    now_dt = datetime.now()
    year_val = str(now_dt.year)
    header_title = title or f"العمليات والصيانة - الشمال - الأعطال - {year_val}"

    # حساب إجمالي المبالغ
    total_amt = 0.0
    for r in rows:
        val = r.get("final_value") or r.get("items_value") or 0
        try:
            total_amt += float(str(val).replace(",", "").strip() or 0)
        except Exception:
            pass

    meta_lines = [
        "التبويب: العمليات والصيانة",
        "التبويب الفرعي: الأعطال",
    ]
    if filters:
        for f in filters:
            if f:
                meta_lines.append(str(f))

    summary_lines = [
        f"إجمالي المبلغ: {total_amt:,.2f} ر.س",
        f"عدد السجلات: {len(rows)}",
    ]

    header_row = brand.apply_brand_header(
        ws,
        title=header_title,
        ncol=ncol,
        meta_lines=meta_lines,
        summary_lines=summary_lines,
    )
    brand.write_header_row(ws, EXPORT_HEADERS, header_row, widths=_COL_WIDTHS)

    start = header_row + 1
    for offset, t in enumerate(rows):
        r = start + offset
        for col, field in enumerate(EXPORT_FIELDS, start=1):
            val = t.get(field)
            if field in _DATE_FIELDS:
                val = _format_date(val) if val else ""
            elif field in _TIME_FIELDS:
                val = _format_time(val) if val else ""
            ws.cell(row=r, column=col, value=val if val is not None else "")

    end = start + len(rows) - 1 if rows else header_row
    if rows:
        brand.style_data_rows(ws, start_row=start, end_row=end, ncol=ncol)
        ws.auto_filter.ref = f"A{header_row}:{get_column_letter(ncol)}{end}"

    return brand.save_workbook_bytes(wb)


def import_tickets_from_excel(file_storage) -> dict:
    wb = load_workbook(file_storage, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {"ok": 0, "updated": 0, "errors": ["الملف فارغ"]}

    header_idx, mapping = _find_header(rows)
    if header_idx < 0 or "ticket_no" not in mapping.values():
        return {"ok": 0, "updated": 0, "errors": ["لم يُعثر على عمود رقم العطل"]}

    inv = {v: k for k, v in mapping.items()}
    created = updated = 0
    errors: list[str] = []
    conn = db.connect()

    for i, row in enumerate(rows[header_idx + 1 :], start=header_idx + 2):
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue
        ticket_no = _cell(row, inv.get("ticket_no"), "ticket_no")
        if not ticket_no:
            errors.append(f"صف {i}: رقم العطل فارغ")
            continue
        # تجاهل صف المثال الافتراضي إن تُرك كما هو دون تعديل جاد — لا نمنع الاستيراد

        data = {}
        for field in TICKET_FIELDS:
            if field == "ticket_no":
                data[field] = ticket_no
                continue
            data[field] = _cell(row, inv.get(field), field) if field in inv else ""

        if data.get("items_value") == "":
            data["items_value"] = None
        if not data.get("status"):
            data["status"] = "تم الإسناد"
        data["status"] = db.normalize_ticket_status(data.get("status"))
        if not (data.get("rekaz_code") or "").strip():
            existing_code = conn.execute(
                "SELECT rekaz_code FROM tickets WHERE ticket_no=?",
                (ticket_no,),
            ).fetchone()
            if existing_code and existing_code["rekaz_code"]:
                data["rekaz_code"] = existing_code["rekaz_code"]
            else:
                data["rekaz_code"] = db.next_series_code("er", conn)

        try:
            existing = conn.execute(
                "SELECT id FROM tickets WHERE ticket_no=?",
                (ticket_no,),
            ).fetchone()
            cols = ", ".join(TICKET_FIELDS)
            if existing:
                sets = ", ".join([f"{f}=?" for f in TICKET_FIELDS])
                conn.execute(
                    f"UPDATE tickets SET {sets}, updated_at=CURRENT_TIMESTAMP WHERE id=?",
                    [data[f] for f in TICKET_FIELDS] + [existing["id"]],
                )
                updated += 1
            else:
                placeholders = ", ".join(["?"] * len(TICKET_FIELDS))
                conn.execute(
                    f"INSERT INTO tickets({cols}) VALUES ({placeholders})",
                    [data[f] for f in TICKET_FIELDS],
                )
                created += 1
        except Exception as exc:
            errors.append(f"صف {i} ({ticket_no}): {exc}")

    conn.commit()
    conn.close()
    return {"ok": created, "updated": updated, "errors": errors}
