"""استيراد بنود العقد الموحد من Excel — دليل BOQ النشط للعقود."""

from __future__ import annotations

from datetime import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins

from webapp import db
from webapp import excel_brand as brand

# رؤوس ثنائية اللغة — الترتيب مطابق لقالب العقد
BOQ_HEADERS = [
    "رقم البند\nItem",
    "نوع البند\nLine type",
    "التوصيف المختصر\nShort Description",
    "التوصيف الكامل\nLong Description",
    "الكمية\nQty",
    "الوحدة\nUOM",
    "سعر الوحدة\nUnit Price",
    "العملة\nCurrency",
    "الاجمالي\nTOTAL",
    "نوع الدفع\nPayment Type",
]

# أعمدة الكمية والإجمالي (1-based) — تظليل رمادي فاتح
_QTY_COL = 5
_TOTAL_COL = 9

_BLACK = PatternFill("solid", fgColor="000000")
_WHITE_BOLD = Font(name="Arial", size=10, bold=True, color="FFFFFF")
_GRAY = PatternFill("solid", fgColor="D9D9D9")
_DATA_FONT = Font(name="Arial", size=10, color="1A1814")
_THIN = Border(
    left=Side(style="thin", color="000000"),
    right=Side(style="thin", color="000000"),
    top=Side(style="thin", color="000000"),
    bottom=Side(style="thin", color="000000"),
)
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_RIGHT = Alignment(horizontal="right", vertical="center", wrap_text=True)

_COL_WIDTHS = {
    1: 14,   # Item
    2: 14,   # Line type
    3: 28,   # Short Description
    4: 55,   # Long Description
    5: 12,   # Qty
    6: 10,   # UOM
    7: 14,   # Unit Price
    8: 11,   # Currency
    9: 14,   # TOTAL
    10: 14,  # Payment Type
}

# مرادفات مرنة للرؤوس (عربي + إنجليزي + صيغ قديمة)
_ALIASES = {
    # item_no
    "رقم البند": "item_no",
    "كود البند": "item_no",
    "رقم البند بالعقد": "item_no",
    "بند": "item_no",
    "item": "item_no",
    "item_no": "item_no",
    "item no": "item_no",
    "item code": "item_no",
    "code": "item_no",
    # line_type
    "نوع البند": "line_type",
    "line type": "line_type",
    "line_type": "line_type",
    "linetype": "line_type",
    # short_desc
    "التوصيف المختصر": "short_desc",
    "الوصف المختصر": "short_desc",
    "وصف مختصر": "short_desc",
    "short description": "short_desc",
    "short_description": "short_desc",
    "short desc": "short_desc",
    "الوصف": "short_desc",
    "وصف البند": "short_desc",
    "بيان البند": "short_desc",
    "البند": "short_desc",
    "description": "short_desc",
    "item description": "short_desc",
    # long_desc
    "التوصيف الكامل": "long_desc",
    "الوصف الكامل": "long_desc",
    "وصف كامل": "long_desc",
    "long description": "long_desc",
    "long_description": "long_desc",
    "long desc": "long_desc",
    # qty
    "الكمية": "qty",
    "كميه": "qty",
    "qty": "qty",
    "quantity": "qty",
    # unit
    "الوحدة": "unit",
    "وحدة القياس": "unit",
    "unit": "unit",
    "uom": "unit",
    # unit_price
    "سعر الوحدة": "unit_price",
    "سعر البند": "unit_price",
    "السعر": "unit_price",
    "سعر": "unit_price",
    "unit_price": "unit_price",
    "unit price": "unit_price",
    "rate": "unit_price",
    # currency
    "العملة": "currency",
    "عمله": "currency",
    "currency": "currency",
    "curr": "currency",
    # amount / TOTAL
    "الاجمالي": "amount",
    "الإجمالي": "amount",
    "المبلغ": "amount",
    "القيمة": "amount",
    "amount": "amount",
    "total": "amount",
    # payment_type
    "نوع الدفع": "payment_type",
    "طريقة الدفع": "payment_type",
    "payment type": "payment_type",
    "payment_type": "payment_type",
    "payment": "payment_type",
    # legacy
    "التصنيف": "category",
    "القسم": "category",
    "category": "category",
    "ملاحظات": "notes",
    "notes": "notes",
}


def _norm_header(value) -> str:
    s = str(value or "").strip().lower()
    s = s.replace("ـ", "").replace("_", " ").replace("/", " ")
    s = s.replace("\r\n", "\n").replace("\r", "\n")
    s = " ".join(part.strip() for part in s.replace("\n", " ").split() if part.strip())
    return s


def _resolve_alias(raw) -> str | None:
    """يطابق رأس عمود ثنائي اللغة أو مرادف واحد."""
    if raw is None:
        return None
    text = str(raw).strip()
    if not text:
        return None
    # تطابق كامل بعد التطبيع
    key = _ALIASES.get(_norm_header(text)) or _ALIASES.get(text)
    if key:
        return key
    # أجزاء مفصولة بسطر جديد أو /
    parts = []
    for chunk in text.replace("/", "\n").split("\n"):
        chunk = chunk.strip()
        if chunk:
            parts.append(chunk)
    for part in parts:
        key = _ALIASES.get(_norm_header(part)) or _ALIASES.get(part)
        if key:
            return key
    # احتواء عبارة معروفة داخل الرأس الطويل
    norm = _norm_header(text)
    # الأطول أولاً لتجنب تطابق جزئي خاطئ (مثل "item" داخل "item description")
    for alias, field in sorted(_ALIASES.items(), key=lambda x: -len(x[0])):
        a = _norm_header(alias)
        if a and (a == norm or f" {a} " in f" {norm} " or norm.startswith(a + " ") or norm.endswith(" " + a)):
            return field
    return None


def _map_headers(row_values) -> dict[int, str]:
    mapping = {}
    for idx, raw in enumerate(row_values):
        key = _resolve_alias(raw)
        if key:
            mapping[idx] = key
    return mapping


def _to_float(val):
    if val is None or val == "":
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace(",", "").replace("٬", "")
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def build_boq_template() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "بنود العقد"
    ws.sheet_view.rightToLeft = True
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)
    ws.print_title_rows = "1:1"

    header_row = 1
    ws.row_dimensions[header_row].height = 36

    for col, label in enumerate(BOQ_HEADERS, 1):
        cell = ws.cell(header_row, col, label)
        cell.fill = _BLACK
        cell.font = _WHITE_BOLD
        cell.alignment = _CENTER
        cell.border = _THIN
        ws.column_dimensions[get_column_letter(col)].width = _COL_WIDTHS.get(col, 14)

    samples = [
        [
            "A1",
            "Description",
            "إنشاء وصيانة شبكات التوزيع",
            "CONSTRUCTION AND MAINTENANCE OF DISTRIBUTION NETWORKS\nإنشاء وصيانة شبكات التوزيع",
            "",
            "",
            "",
            "SAR",
            "",
            "",
        ],
        [
            "10000000",
            "Description",
            "أعمال المساحة ونظم المعلومات الجغرافية",
            "SURVEYING AND GEOGRAPHICAL INFORMATION SYSTEM (GIS) WORKS\nأعمال المساحة ونظم المعلومات الجغرافية",
            "",
            "",
            "",
            "SAR",
            "",
            "",
        ],
        [
            "10100000",
            "Description",
            "أعمال المسح",
            "SURVEYING WORKS\nأعمال المسح — يشمل أجهزة GPS ومحطة كلية (Total Station) حسب متطلبات العقد",
            "",
            "LS",
            "",
            "SAR",
            "",
            "",
        ],
    ]

    for r_i, sample in enumerate(samples):
        row = header_row + 1 + r_i
        ws.row_dimensions[row].height = 48
        for col, val in enumerate(sample, 1):
            cell = ws.cell(row, col, val if val != "" else None)
            cell.font = _DATA_FONT
            cell.alignment = _CENTER if col in (1, 2, 5, 6, 7, 8, 9, 10) else _RIGHT
            cell.border = _THIN
            if col in (_QTY_COL, _TOTAL_COL):
                cell.fill = _GRAY

    # صفوف فارغة إضافية بنفس تظليل Qty/TOTAL
    for extra in range(3):
        row = header_row + 1 + len(samples) + extra
        for col in range(1, len(BOQ_HEADERS) + 1):
            cell = ws.cell(row, col, None)
            cell.border = _THIN
            if col in (_QTY_COL, _TOTAL_COL):
                cell.fill = _GRAY

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(BOQ_HEADERS))}1"

    brand.write_instructions_sheet(
        wb,
        [
            "الأعمدة: رقم البند، نوع البند، التوصيف المختصر، التوصيف الكامل، الكمية، الوحدة، سعر الوحدة، العملة، الاجمالي، نوع الدفع.",
            "يمكن استخدام الرؤوس بالعربية أو الإنجليزية (Item, Line type, Short Description, …).",
            "الكمية وسعر الوحدة يُستخدمان عند اختيار البند داخل العطل؛ الكمية تُدخل على العطل وسعر الوحدة من الدليل.",
            "عند الرفع يُستبدل الدليل النشط بالملف الجديد مع الاحتفاظ بسجل الملفات السابقة.",
        ],
    )
    return brand.save_workbook_bytes(wb)


def import_boq_from_excel(file_storage, *, uploaded_by: str = "") -> dict:
    """يفسّر Excel ويحفظه كدليل بنود عقد نشط. يعيد إحصاءات الاستيراد."""
    wb = load_workbook(file_storage, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header_map = None
    header_row_idx = 0
    for i, row in enumerate(rows_iter, 1):
        mapped = _map_headers(row)
        vals = set(mapped.values())
        if "item_no" in vals and (
            "short_desc" in vals or "long_desc" in vals or "unit_price" in vals or "description" in vals
        ):
            header_map = mapped
            header_row_idx = i
            break
    if not header_map:
        raise ValueError(
            "لم يُعثر على صف رؤوس. المتوقع: رقم البند / Item مع التوصيف أو سعر الوحدة (أو مرادفاتها)."
        )

    field_defaults = (
        "item_no",
        "line_type",
        "short_desc",
        "long_desc",
        "qty",
        "unit",
        "unit_price",
        "currency",
        "amount",
        "payment_type",
        "category",
        "notes",
    )
    items = []
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        data = {field: None for field in field_defaults}
        for idx, field in header_map.items():
            if idx >= len(row):
                continue
            val = row[idx]
            if field in ("unit_price", "amount", "qty"):
                data[field] = _to_float(val)
            else:
                data[field] = str(val).strip() if val is not None and str(val).strip() else ""
        item_no = (data.get("item_no") or "").strip()
        if not item_no:
            continue
        short_desc = (data.get("short_desc") or "").strip()
        long_desc = (data.get("long_desc") or "").strip()
        # الوصف المعروض/المتوافق مع الكود القديم = المختصر ثم الكامل
        description = short_desc or long_desc
        price = data.get("unit_price")
        if price is None and data.get("amount") is not None and not description:
            continue
        items.append(
            {
                "item_no": item_no,
                "line_type": (data.get("line_type") or "").strip(),
                "short_desc": short_desc,
                "long_desc": long_desc,
                "description": description,
                "unit": (data.get("unit") or "").strip(),
                "unit_price": price,
                "currency": (data.get("currency") or "").strip(),
                "amount": data.get("amount"),
                "payment_type": (data.get("payment_type") or "").strip(),
                "category": (data.get("category") or "").strip(),
                "notes": (data.get("notes") or "").strip(),
            }
        )

    if not items:
        raise ValueError("الملف لا يحتوي بنوداً صالحة بعد صف الرؤوس.")

    filename = getattr(file_storage, "filename", None) or "بنود_العقد.xlsx"
    conn = db.connect()
    try:
        db.ensure_schema(conn)
        conn.execute("UPDATE contract_boq_files SET is_active=0 WHERE is_active=1")
        cur = conn.execute(
            """
            INSERT INTO contract_boq_files(filename, uploaded_at, uploaded_by, is_active, item_count, notes)
            VALUES (?,?,?,?,?,?)
            """,
            (
                filename,
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                uploaded_by or "",
                1,
                len(items),
                "",
            ),
        )
        file_id = cur.lastrowid
        conn.executemany(
            """
            INSERT INTO contract_boq_items(
              file_id, item_no, description, short_desc, long_desc, line_type,
              unit, unit_price, currency, amount, payment_type, category, notes
            ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)
            """,
            [
                (
                    file_id,
                    it["item_no"],
                    it["description"],
                    it["short_desc"],
                    it["long_desc"],
                    it["line_type"],
                    it["unit"],
                    it["unit_price"],
                    it["currency"],
                    it["amount"],
                    it["payment_type"],
                    it["category"],
                    it["notes"],
                )
                for it in items
            ],
        )
        # مزامنة جدول boq_itemslegacy كمرآة للدليل النشط
        conn.execute("DELETE FROM boq_items")
        conn.executemany(
            """
            INSERT INTO boq_items(
              item_no, description, short_desc, long_desc, line_type,
              unit, unit_price, currency, payment_type, category, notes
            ) VALUES (?,?,?,?,?,?,?,?,?,?,?)
            """,
            [
                (
                    it["item_no"],
                    it["description"],
                    it["short_desc"],
                    it["long_desc"],
                    it["line_type"],
                    it["unit"],
                    it["unit_price"],
                    it["currency"],
                    it["payment_type"],
                    it["category"],
                    it["notes"],
                )
                for it in items
            ],
        )
        conn.commit()
    finally:
        conn.close()

    return {"ok": len(items), "file_id": file_id, "filename": filename}
