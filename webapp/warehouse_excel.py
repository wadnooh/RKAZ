"""استيراد وتصدير مواد وحركات المستودع من/إلى Excel."""

from __future__ import annotations

from datetime import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from webapp import db
from webapp import excel_brand as brand

ITEM_HEADERS = [
    "رقم المادة",
    "اسم المادة",
    "الوحدة",
    "التصنيف",
    "حد أدنى",
    "ملاحظات",
    "رصيد افتتاحي",
]

TX_HEADERS = [
    "رقم السند",
    "تاريخ الحركة",
    "نوع الحركة",
    "رقم المادة",
    "اسم المادة",
    "الوحدة",
    "الكمية",
    "المستلم",
    "المسلم",
    "رقم العطل",
    "المنطقة",
    "ملاحظات",
]

# مرادفات رؤوس الأعمدة
_ITEM_ALIASES = {
    "رقم المادة": "item_no",
    "كود المادة": "item_no",
    "رقم الصنف": "item_no",
    "item_no": "item_no",
    "item no": "item_no",
    "اسم المادة": "item_name",
    "اسم الصنف": "item_name",
    "المادة": "item_name",
    "item_name": "item_name",
    "الوحدة": "unit",
    "unit": "unit",
    "التصنيف": "category",
    "category": "category",
    "حد أدنى": "min_qty",
    "الحد الأدنى": "min_qty",
    "min_qty": "min_qty",
    "ملاحظات": "notes",
    "notes": "notes",
    "رصيد افتتاحي": "opening_qty",
    "الرصيد الافتتاحي": "opening_qty",
    "opening": "opening_qty",
    "opening_qty": "opening_qty",
}

_TX_ALIASES = {
    "رقم السند": "voucher_no",
    "السند": "voucher_no",
    "voucher_no": "voucher_no",
    "تاريخ الحركة": "tx_date",
    "التاريخ": "tx_date",
    "tx_date": "tx_date",
    "نوع الحركة": "tx_type",
    "الحركة": "tx_type",
    "tx_type": "tx_type",
    "رقم المادة": "item_no",
    "كود المادة": "item_no",
    "item_no": "item_no",
    "اسم المادة": "item_name",
    "item_name": "item_name",
    "الوحدة": "unit",
    "unit": "unit",
    "الكمية": "qty",
    "qty": "qty",
    "المستلم / المسلم": "recipient",
    "المستلم": "recipient",
    "recipient": "recipient",
    "المسلم": "sender",
    "sender": "sender",
    "رقم العطل": "ticket_no",
    "رقم البلاغ": "ticket_no",  # توافق مع القوالب القديمة
    "البلاغ": "ticket_no",
    "ticket_no": "ticket_no",
    "fault no": "ticket_no",
    "fault_no": "ticket_no",
    "المنطقة": "region",
    "region": "region",
    "ملاحظات": "notes",
    "notes": "notes",
}


def _norm_header(value) -> str:
    return str(value or "").strip().lower().replace("ـ", "")


def _map_headers(row_values, aliases: dict) -> dict[int, str]:
    mapping = {}
    for idx, raw in enumerate(row_values):
        key = str(raw or "").strip()
        field = aliases.get(key) or aliases.get(key.lower())
        if not field:
            field = aliases.get(_norm_header(key))
        if field:
            mapping[idx] = field
    return mapping


def _cell(row, idx):
    if idx is None or idx >= len(row):
        return ""
    val = row[idx]
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    return str(val).strip()


def _to_float(val):
    if val is None or val == "":
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace(",", "")
    if not s:
        return None
    try:
        return float(s)
    except Exception:
        return None


def _find_header(rows, aliases: dict, required_any: set[str], max_scan: int = 20) -> tuple[int, dict[int, str]]:
    for i, row in enumerate(rows[:max_scan]):
        if not row:
            continue
        mapping = _map_headers(row, aliases)
        if required_any & set(mapping.values()):
            return i, mapping
    return -1, {}


def build_items_template() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "المواد"
    ncol = len(ITEM_HEADERS)
    header_row = brand.apply_brand_header(ws, title="قالب استيراد مواد المستودع", ncol=ncol)
    brand.write_header_row(ws, ITEM_HEADERS, header_row)
    # صفوف أمثلة للقالب
    samples = [
        ["908111006", "CABLE, PWR, 600V/1KV, AL, 4C, 185MM2, XLPE", "KM", "CABLE, PWR, 600V/1KV, AL, 4C, 185MM2, XLPE", 50, "", ""],
        ["908202053", "ROD,GRD,CUWLD STL,16MM DIA,1200MM LG", "EA", "ROD,GRD,CUWLD STL,16MM DIA,1200MM LG", 50, "", ""],
    ]
    for offset, sample in enumerate(samples):
        r = header_row + 1 + offset
        for col, val in enumerate(sample, start=1):
            ws.cell(row=r, column=col, value=val)
    end = header_row + len(samples)
    brand.style_data_rows(ws, start_row=header_row + 1, end_row=end, ncol=ncol)
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(ncol)}{end}"
    brand.write_instructions_sheet(
        wb,
        [
            "ارفع ملف المواد من شاشة أصناف المستودع.",
            "عمود رصيد افتتاحي اختياري — يُنشئ حركة وارد (رصيد افتتاحي) مربوطة بالمادة.",
            "رقم المادة مطلوب وفريد — عند التكرار يتم تحديث بيانات المادة.",
            "لا تحذف صف رؤوس الأعمدة.",
        ],
    )
    return brand.save_workbook_bytes(wb)


def build_tx_template() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "الحركات"
    ncol = len(TX_HEADERS)
    header_row = brand.apply_brand_header(ws, title="قالب استيراد حركات المستودع", ncol=ncol)
    brand.write_header_row(ws, TX_HEADERS, header_row)
    today = datetime.now().strftime("%Y-%m-%d")
    samples = [
        ["V-001", today, "وارد من الكهرباء", "M-001", "كيبل 4×16", "متر", 100, "المستودع", "", "خريص", "مثال وارد"],
        ["V-002", today, "منصرف للمقاول", "M-001", "كيبل 4×16", "متر", 25, "فرقة 1", "T-100", "خريص", "ربط بالمعاملة/العطل"],
    ]
    for offset, sample in enumerate(samples):
        r = header_row + 1 + offset
        for col, val in enumerate(sample, start=1):
            ws.cell(row=r, column=col, value=val)
    end = header_row + len(samples)
    brand.style_data_rows(ws, start_row=header_row + 1, end_row=end, ncol=ncol)
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(ncol)}{end}"
    brand.write_instructions_sheet(
        wb,
        [
            "أنواع الحركة المقترحة: وارد من الكهرباء | منصرف للمقاول | إرجاع للمجمعة | وارد من موقع العمل | رصيد افتتاحي",
            "رقم المادة يجب أن يكون موجوداً في أصناف المستودع (أو يُنشأ تلقائياً إن وُجد الاسم).",
            "رقم العطل يربط الحركة بمعاملة العطل.",
            "لا تحذف صف رؤوس الأعمدة.",
        ],
    )
    return brand.save_workbook_bytes(wb)


def _pick_sheet(wb, preferred: tuple[str, ...]):
    for name in preferred:
        if name in wb.sheetnames:
            return wb[name]
    return wb.active


def import_items_from_excel(file_storage) -> dict:
    wb = load_workbook(file_storage, data_only=True)
    ws = _pick_sheet(wb, ("المواد", "مواد", "Items", "items"))
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {"ok": 0, "updated": 0, "opening": 0, "errors": ["الملف فارغ"]}

    header_idx, mapping = _find_header(rows, _ITEM_ALIASES, {"item_no", "item_name"})
    if header_idx < 0:
        return {"ok": 0, "updated": 0, "opening": 0, "errors": ["لم يُعثر على أعمدة رقم/اسم المادة"]}

    inv = {v: k for k, v in mapping.items()}
    created = updated = opening = 0
    errors = []
    conn = db.connect()
    today = datetime.now().strftime("%Y-%m-%d")

    for i, row in enumerate(rows[header_idx + 1 :], start=header_idx + 2):
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue
        item_no = _cell(row, inv.get("item_no"))
        item_name = _cell(row, inv.get("item_name"))
        if not item_no and not item_name:
            continue
        if not item_no:
            item_no = f"AUTO-{i}"
        unit = _cell(row, inv.get("unit")) or "عدد"
        category = _cell(row, inv.get("category")) or "مواد كهربائية"
        min_qty = _to_float(_cell(row, inv.get("min_qty")) if "min_qty" in inv else None) or 0
        notes = _cell(row, inv.get("notes"))
        open_qty = _to_float(_cell(row, inv.get("opening_qty")) if "opening_qty" in inv else None)

        existing = conn.execute(
            "SELECT id FROM warehouse_items WHERE lower(item_no)=lower(?)",
            (item_no,),
        ).fetchone()
        if existing:
            conn.execute(
                """
                UPDATE warehouse_items
                SET item_name=?, unit=?, category=?, min_qty=?, notes=?
                WHERE id=?
                """,
                (item_name or item_no, unit, category, min_qty, notes, existing["id"]),
            )
            updated += 1
        else:
            conn.execute(
                """
                INSERT INTO warehouse_items(item_no, item_name, unit, category, min_qty, notes)
                VALUES (?,?,?,?,?,?)
                """,
                (item_no, item_name or item_no, unit, category, min_qty, notes),
            )
            created += 1

        if open_qty and open_qty > 0:
            # تجنب تكرار رصيد افتتاحي لنفس المادة
            already = conn.execute(
                """
                SELECT id FROM warehouse_tx
                WHERE lower(item_no)=lower(?) AND tx_type=? AND qty=?
                LIMIT 1
                """,
                (item_no, "رصيد افتتاحي", open_qty),
            ).fetchone()
            if not already:
                conn.execute(
                    """
                    INSERT INTO warehouse_tx(
                        voucher_no, tx_date, tx_type, item_no, item_name, unit, qty,
                        recipient, sender, ticket_no, region, notes
                    ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
                    """,
                    (
                        f"OPEN-{item_no}",
                        today,
                        "رصيد افتتاحي",
                        item_no,
                        item_name or item_no,
                        unit,
                        open_qty,
                        "استيراد Excel",
                        "",
                        "",
                        "",
                        "رصيد افتتاحي من استيراد المواد",
                    ),
                )
                opening += 1

    conn.commit()
    conn.close()
    return {"ok": created, "updated": updated, "opening": opening, "errors": errors}


def import_tx_from_excel(file_storage) -> dict:
    wb = load_workbook(file_storage, data_only=True)
    ws = _pick_sheet(wb, ("الحركات", "حركات", "Transactions", "tx"))
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {"ok": 0, "linked": 0, "errors": ["الملف فارغ"]}

    header_idx, mapping = _find_header(rows, _TX_ALIASES, {"item_no", "item_name"})
    if header_idx < 0:
        return {"ok": 0, "linked": 0, "errors": ["لم يُعثر على عمود المادة"]}
    if "qty" not in mapping.values():
        return {"ok": 0, "linked": 0, "errors": ["لم يُعثر على عمود الكمية"]}

    inv = {v: k for k, v in mapping.items()}
    created = linked = 0
    errors = []
    conn = db.connect()
    today = datetime.now().strftime("%Y-%m-%d")

    for i, row in enumerate(rows[header_idx + 1 :], start=header_idx + 2):
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue
        item_no = _cell(row, inv.get("item_no"))
        item_name = _cell(row, inv.get("item_name"))
        qty = _to_float(row[inv["qty"]] if "qty" in inv else None)
        if qty is None or qty == 0:
            errors.append(f"صف {i}: كمية غير صالحة")
            continue
        if not item_no and not item_name:
            errors.append(f"صف {i}: بدون مادة")
            continue

        # ربط / إنشاء الصنف
        item = None
        if item_no:
            item = conn.execute(
                "SELECT * FROM warehouse_items WHERE lower(item_no)=lower(?)",
                (item_no,),
            ).fetchone()
        if not item and item_name:
            item = conn.execute(
                "SELECT * FROM warehouse_items WHERE lower(item_name)=lower(?)",
                (item_name,),
            ).fetchone()
        if not item:
            item_no = item_no or f"AUTO-TX-{i}"
            unit = _cell(row, inv.get("unit")) or "عدد"
            conn.execute(
                """
                INSERT INTO warehouse_items(item_no, item_name, unit, category, min_qty, notes)
                VALUES (?,?,?,?,?,?)
                """,
                (item_no, item_name or item_no, unit, "مواد كهربائية", 0, "أُنشئ من استيراد الحركات"),
            )
            item = conn.execute(
                "SELECT * FROM warehouse_items WHERE lower(item_no)=lower(?)",
                (item_no,),
            ).fetchone()
        else:
            item_no = item["item_no"]
            if not item_name:
                item_name = item["item_name"]

        unit = _cell(row, inv.get("unit")) or (item["unit"] if item else "عدد")
        tx_type = _cell(row, inv.get("tx_type")) or "وارد من الكهرباء"
        ticket_no = _cell(row, inv.get("ticket_no"))
        voucher_no = _cell(row, inv.get("voucher_no")) or f"TX-{today}-{i}"
        tx_date = today
        if inv.get("tx_date") is not None:
            raw_date = row[inv["tx_date"]]
            if isinstance(raw_date, datetime):
                tx_date = raw_date.strftime("%Y-%m-%d")
            else:
                tx_date = _cell(row, inv.get("tx_date")) or today

        conn.execute(
            """
            INSERT INTO warehouse_tx(
                voucher_no, tx_date, tx_type, item_no, item_name, unit, qty,
                recipient, sender, ticket_no, region, notes
            ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
            """,
            (
                voucher_no,
                tx_date,
                tx_type,
                item_no,
                item_name or (item["item_name"] if item else item_no),
                unit,
                qty,
                _cell(row, inv.get("recipient")),
                _cell(row, inv.get("sender")),
                ticket_no,
                _cell(row, inv.get("region")),
                _cell(row, inv.get("notes")),
            ),
        )
        created += 1
        if ticket_no:
            linked += 1

    conn.commit()
    conn.close()
    return {"ok": created, "linked": linked, "errors": errors[:30]}
