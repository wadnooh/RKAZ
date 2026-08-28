"""تنسيق موحّد واحترافي لملفات Excel — ترويسة، شعارات، فلاتر وإجماليات بنمط ركاز."""

from __future__ import annotations

import io
from datetime import datetime
from pathlib import Path

from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

BRAND_DIR = Path(__file__).resolve().parent / "static" / "brand"
LOGO_REKAZ = BRAND_DIR / "rekaz.png"
LOGO_SEC = BRAND_DIR / "sec.jpg"

COMPANY_NAME = "شركة ركاز الإنجاز للمقاولات"
OFFICE_NAME = "مكتب خدمات خريص"

# ألوان الترويسة والجداول
COLOR_TITLE_BG = "D9D9D9"      # رمادي / فضي للبانر الرئيسي
COLOR_TITLE_FG = "1F497D"      # كحلي داكن للنصوص الرئيسية
COLOR_META_BG = "F2F2F9"       # خلفية ليلكية / ثلجية ناعمة لبطاقات الفلاتر
COLOR_META_FG = "1F497D"       # نص كحلي للفلاتر
COLOR_SUMMARY_BG = "FFFDF8"    # خلفية كريمية ذهبية خفيفة للإجماليات
COLOR_SUMMARY_FG = "002060"    # كحلي ملكي للإجماليات
COLOR_HEADER_BG = "1F4E79"     # كحلي داكن لرؤوس الأعمدة
COLOR_HEADER_FG = "FFFFFF"     # أبيض لرؤوس الأعمدة
COLOR_ALT_ROW = "F8FAFC"       # تظليل تبادلي خفيف
COLOR_BORDER = "D0D5DD"        # إطار رمادي ناعم

_THIN = Border(
    left=Side(style="thin", color=COLOR_BORDER),
    right=Side(style="thin", color=COLOR_BORDER),
    top=Side(style="thin", color=COLOR_BORDER),
    bottom=Side(style="thin", color=COLOR_BORDER),
)

_HEADER_BORDER = Border(
    left=Side(style="thin", color="1A4064"),
    right=Side(style="thin", color="1A4064"),
    top=Side(style="thin", color="1A4064"),
    bottom=Side(style="medium", color="0D2640"),
)


def logo_file() -> Path | None:
    return LOGO_REKAZ if LOGO_REKAZ.is_file() else None


def sec_logo_file() -> Path | None:
    return LOGO_SEC if LOGO_SEC.is_file() else None


def _add_logo(
    ws: Worksheet,
    path: Path | None,
    anchor: str,
    max_height: int = 46,
    max_width: int = 120,
) -> None:
    if not path or not path.is_file():
        return
    try:
        img = XLImage(str(path))
        if img.height and img.height > max_height:
            ratio = max_height / float(img.height)
            img.height = max_height
            img.width = int(img.width * ratio)
        if img.width and img.width > max_width:
            ratio = max_width / float(img.width)
            img.width = max_width
            img.height = int(img.height * ratio)
        ws.add_image(img, anchor)
    except Exception:
        pass


def apply_brand_header(
    ws: Worksheet,
    *,
    title: str,
    ncol: int,
    subtitle: str | None = None,
    as_of: datetime | None = None,
    section: str | None = None,
    sub_section: str | None = None,
    category: str | None = None,
    area: str | None = None,
    supplier: str | None = None,
    period: str | None = None,
    search_q: str | None = None,
    meta_lines: list[str] | None = None,
    summary_lines: list[str] | None = None,
) -> int:
    """
    يبني ترويسة احترافية كاملة متطابقة مع التصميم المطلوب:
    1. صف البانر الفضي الرئيسي (العنوان + الشعارات)
    2. صف التاريخ
    3. سطور الفلاتر والبيانات الوصفية
    4. سطور الإجماليات والملخص
    5. صف فاصل فارغ
    يعيد رقم صف رؤوس الأعمدة.
    """
    as_of = as_of or datetime.now()
    ncol = max(ncol, 4)
    last_col = get_column_letter(ncol)
    ws.sheet_view.rightToLeft = True

    # 1. الصف الأول: البانر الفضي الرئيسي
    ws.row_dimensions[1].height = 44
    ws.merge_cells(f"A1:{last_col}1")
    title_cell = ws["A1"]
    title_cell.value = title
    title_cell.font = Font(name="Arial", size=15, bold=True, color=COLOR_TITLE_FG)
    title_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    title_cell.fill = PatternFill("solid", fgColor=COLOR_TITLE_BG)
    title_cell.border = _THIN

    # إضافة الشعارات (ركاز في A1 وكهرباء في العمود الأخير)
    _add_logo(ws, LOGO_REKAZ, "A1", max_height=42, max_width=110)
    if ncol >= 5:
        _add_logo(ws, LOGO_SEC, f"{last_col}1", max_height=40, max_width=110)

    # 2. الصف الثاني: التاريخ
    ws.row_dimensions[2].height = 22
    ws.merge_cells(f"A2:{last_col}2")
    date_cell = ws["A2"]
    date_cell.value = f"التاريخ: {as_of.strftime('%d-%m-%Y')}"
    date_cell.font = Font(name="Arial", size=11, bold=True, color=COLOR_TITLE_FG)
    date_cell.alignment = Alignment(horizontal="center", vertical="center")
    date_cell.fill = PatternFill("solid", fgColor="FFFFFF")
    date_cell.border = _THIN

    # 3. سطور الفلاتر والبيانات الوصفية
    lines: list[str] = []
    if meta_lines is not None:
        lines = [line for line in meta_lines if line]
    else:
        if section or sub_section or category or area or supplier or period or search_q:
            if section:
                lines.append(f"التبويب: {section}")
            if sub_section:
                lines.append(f"التبويب الفرعي: {sub_section}")
            if category:
                lines.append(f"النوع: {category}")
            if area:
                lines.append(f"المنطقة: {area}")
            if supplier:
                lines.append(f"المورد / المندوب: {supplier}")
            if period:
                lines.append(f"الفترة: {period}")
            if search_q:
                lines.append(f"بحث: {search_q}")

    current_row = 3
    for line in lines:
        ws.row_dimensions[current_row].height = 21
        ws.merge_cells(f"A{current_row}:{last_col}{current_row}")
        c = ws[f"A{current_row}"]
        c.value = line
        c.font = Font(name="Arial", size=11, bold=True, color=COLOR_META_FG)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.fill = PatternFill("solid", fgColor=COLOR_META_BG)
        c.border = _THIN
        current_row += 1

    # 4. سطور الإجماليات والملخص
    if summary_lines:
        for sline in summary_lines:
            if not sline:
                continue
            ws.row_dimensions[current_row].height = 24
            ws.merge_cells(f"A{current_row}:{last_col}{current_row}")
            sc = ws[f"A{current_row}"]
            sc.value = sline
            sc.font = Font(name="Arial", size=12, bold=True, color=COLOR_SUMMARY_FG)
            sc.alignment = Alignment(horizontal="center", vertical="center")
            sc.fill = PatternFill("solid", fgColor=COLOR_SUMMARY_BG)
            sc.border = _THIN
            current_row += 1

    # 5. صف فاصل فارغ
    ws.row_dimensions[current_row].height = 8
    current_row += 1

    return current_row


def write_header_row(
    ws: Worksheet,
    headers: list[str],
    row: int,
    *,
    widths: dict[str, float] | None = None,
) -> None:
    fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
    font = Font(name="Arial", size=11, bold=True, color=COLOR_HEADER_FG)
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 28

    for idx, label in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=idx, value=label)
        cell.fill = fill
        cell.font = font
        cell.alignment = align
        cell.border = _HEADER_BORDER
        letter = get_column_letter(idx)
        if widths and label in widths:
            ws.column_dimensions[letter].width = widths[label]
        else:
            ws.column_dimensions[letter].width = max(14, min(32, len(label) * 1.6 + 6))

    ws.freeze_panes = f"A{row + 1}"
    ws.auto_filter.ref = f"A{row}:{get_column_letter(len(headers))}{row}"


def style_data_rows(
    ws: Worksheet,
    *,
    start_row: int,
    end_row: int,
    ncol: int,
) -> None:
    if end_row < start_row:
        return
    alt = PatternFill("solid", fgColor=COLOR_ALT_ROW)
    align_center = Alignment(horizontal="center", vertical="center")
    font = Font(name="Arial", size=10, color="1A202C")

    for r in range(start_row, end_row + 1):
        ws.row_dimensions[r].height = 22
        for c in range(1, ncol + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = font
            cell.border = _THIN
            if (r - start_row) % 2 == 1:
                cell.fill = alt

            val = cell.value
            if isinstance(val, (int, float)):
                cell.alignment = align_center
                if isinstance(val, float):
                    cell.number_format = "#,##0.00"
                else:
                    cell.number_format = "#,##0"
            elif isinstance(val, str) and val.strip().replace(",", "").replace(".", "").isdigit():
                cell.alignment = align_center
            else:
                cell.alignment = align_center


def write_instructions_sheet(wb, lines: list[str], title: str = "تعليمات") -> None:
    tip = wb.create_sheet(title)
    tip.sheet_view.rightToLeft = True
    tip.column_dimensions["A"].width = 72
    tip["A1"] = f"{COMPANY_NAME} — {OFFICE_NAME}"
    tip["A1"].font = Font(name="Arial", size=12, bold=True, color=COLOR_TITLE_FG)
    tip["A2"] = "إرشادات تعبئة القالب"
    tip["A2"].font = Font(name="Arial", size=11, bold=True, color=COLOR_HEADER_BG)
    tip.row_dimensions[1].height = 22
    tip.row_dimensions[2].height = 20
    for i, line in enumerate(lines, start=4):
        cell = tip.cell(row=i, column=1, value=f"• {line}")
        cell.font = Font(name="Arial", size=10, color="1A202C")
        cell.alignment = Alignment(horizontal="right", wrap_text=True)
        tip.row_dimensions[i].height = 18


def save_workbook_bytes(wb) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()

